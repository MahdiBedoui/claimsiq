"""Report generation - Excel and PDF."""
import os
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from core.database import query
from config import REPORTS_DIR, COMPANY_NAME

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CURRENCY = '$#,##0.00'


def _hdr(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style(ws, r1, r2, ncol, cur_cols=None):
    cur_cols = cur_cols or []
    for r in range(r1, r2 + 1):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            if c in cur_cols:
                cell.number_format = CURRENCY
            if r % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def _autowidth(ws):
    for col in ws.columns:
        ml = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ml + 2, 12), 30)


def _get_summary():
    by_type = query("SELECT * FROM vw_claims_by_type")
    by_status = query("SELECT * FROM vw_claims_by_status")
    monthly = query("SELECT * FROM vw_monthly_trend")
    by_state = query("SELECT * FROM vw_by_state")
    totals = query("SELECT COUNT(*) as cnt, SUM(claim_amount) as total, AVG(claim_amount) as avg FROM claims")
    t = totals[0] if totals else {"cnt": 0, "total": 0, "avg": 0}
    return {
        "total_claims": t["cnt"], "total_claimed": t["total"], "avg_claim": t["avg"],
        "by_type": by_type, "by_status": by_status, "monthly": monthly, "by_state": by_state
    }


def generate_excel(report_type: str) -> str:
    os.makedirs(REPORTS_DIR, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")
    wb = Workbook()
    s = _get_summary()

    if report_type == "daily":
        ws = wb.active
        ws.title = "Daily Summary"
        claims = query("SELECT * FROM claims ORDER BY claim_date DESC LIMIT 100")
        ws.cell(row=1, column=1, value=f"{COMPANY_NAME} — Daily Claims Report").font = Font(bold=True, size=14, color="2F5496")
        ws.cell(row=2, column=1, value=f"Date: {today}").font = Font(italic=True, color="808080")
        headers = ["Claim ID","Policy ID","Claimant","Type","Amount","Status","Date","Provider"]
        r = 4
        for i, h in enumerate(headers, 1):
            ws.cell(row=r, column=i, value=h)
        _hdr(ws, r, len(headers))
        dr = r + 1
        for cl in claims:
            r += 1
            for i, k in enumerate(["claim_id","policy_id","claimant_name","claim_type","claim_amount","status","claim_date","provider_name"], 1):
                ws.cell(row=r, column=i, value=cl.get(k))
        _style(ws, dr, r, len(headers), cur_cols=[5])
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws.conditional_formatting.add(f"F{dr}:F{r}", CellIsRule(operator="equal", formula=['"Denied"'], fill=red_fill))
        ws.conditional_formatting.add(f"F{dr}:F{r}", CellIsRule(operator="equal", formula=['"Closed"'], fill=green_fill))
        _autowidth(ws)

    elif report_type == "weekly":
        ws = wb.active
        ws.title = "Weekly Summary"
        ws.cell(row=1, column=1, value=f"{COMPANY_NAME} — Weekly Claims Report").font = Font(bold=True, size=14, color="2F5496")
        ws.cell(row=2, column=1, value=f"Week ending: {today}").font = Font(italic=True, color="808080")
        hdrs = ["Claim Type","Total Claims","Total Claimed","Avg Claimed","Total Approved","Approval %","Avg Days"]
        r = 4
        for i, h in enumerate(hdrs, 1):
            ws.cell(row=r, column=i, value=h)
        _hdr(ws, r, len(hdrs))
        dr = r + 1
        for item in s["by_type"]:
            r += 1
            ws.cell(row=r, column=1, value=item["claim_type"])
            ws.cell(row=r, column=2, value=item["total_claims"])
            ws.cell(row=r, column=3, value=item["total_claimed"])
            ws.cell(row=r, column=4, value=item["avg_claimed"])
            ws.cell(row=r, column=5, value=item["total_approved"])
            ws.cell(row=r, column=6, value=(item["approval_rate_pct"] or 0) / 100)
            ws.cell(row=r, column=7, value=round(item["avg_processing_days"] or 0, 1))
        _style(ws, dr, r, len(hdrs), cur_cols=[3,4,5])
        chart = BarChart()
        chart.type = "col"
        chart.title = "Total Claimed by Type"
        chart.style = 10
        chart.width = 18
        chart.height = 10
        cats = Reference(ws, min_col=1, min_row=dr, max_row=r)
        vals = Reference(ws, min_col=3, min_row=dr-1, max_row=r)
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{r+3}")
        _autowidth(ws)

        # Status sheet
        ws2 = wb.create_sheet("Status")
        ws2.cell(row=1, column=1, value="Claims by Status").font = Font(bold=True, size=14, color="2F5496")
        hdrs2 = ["Status","Count","Total Amount","Avg Amount"]
        for i, h in enumerate(hdrs2, 1):
            ws2.cell(row=3, column=i, value=h)
        _hdr(ws2, 3, len(hdrs2))
        r2 = 4
        for item in s["by_status"]:
            ws2.cell(row=r2, column=1, value=item["status"])
            ws2.cell(row=r2, column=2, value=item["total_claims"])
            ws2.cell(row=r2, column=3, value=item["total_amount"])
            ws2.cell(row=r2, column=4, value=item["avg_amount"])
            r2 += 1
        _style(ws2, 4, r2-1, len(hdrs2), cur_cols=[3,4])
        pie = PieChart()
        pie.title = "Status Distribution"
        pie.style = 10
        pie.width = 14
        pie.height = 10
        cats = Reference(ws2, min_col=1, min_row=4, max_row=r2-1)
        vals = Reference(ws2, min_col=2, min_row=3, max_row=r2-1)
        pie.add_data(vals, titles_from_data=True)
        pie.set_categories(cats)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        ws2.add_chart(pie, f"A{r2+2}")
        _autowidth(ws2)

    elif report_type == "monthly":
        ws = wb.active
        ws.title = "Monthly Trend"
        ws.cell(row=1, column=1, value=f"{COMPANY_NAME} — Monthly Trend Report").font = Font(bold=True, size=14, color="2F5496")
        ws.cell(row=2, column=1, value=f"Generated: {today}").font = Font(italic=True, color="808080")
        hdrs = ["Month","Claims","Total Amount","Avg Amount","Denied","Closed"]
        r = 4
        for i, h in enumerate(hdrs, 1):
            ws.cell(row=r, column=i, value=h)
        _hdr(ws, r, len(hdrs))
        dr = r + 1
        for item in s["monthly"]:
            r += 1
            ws.cell(row=r, column=1, value=item["month"])
            ws.cell(row=r, column=2, value=item["total_claims"])
            ws.cell(row=r, column=3, value=item["total_amount"])
            ws.cell(row=r, column=4, value=item["avg_amount"])
            ws.cell(row=r, column=5, value=item["denied"])
            ws.cell(row=r, column=6, value=item["closed"])
        _style(ws, dr, r, len(hdrs), cur_cols=[3,4])
        line = LineChart()
        line.title = "Monthly Claims Trend"
        line.style = 10
        line.width = 20
        line.height = 12
        cats = Reference(ws, min_col=1, min_row=dr, max_row=r)
        vals = Reference(ws, min_col=2, min_row=dr-1, max_row=r)
        line.add_data(vals, titles_from_data=True)
        line.set_categories(cats)
        ws.add_chart(line, f"A{r+3}")
        _autowidth(ws)

    fname = f"{report_type}_report_{today}.xlsx"
    fpath = os.path.join(REPORTS_DIR, fname)
    wb.save(fpath)
    return fpath


def generate_pdf(report_type: str) -> str:
    os.makedirs(REPORTS_DIR, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")
    fname = f"{report_type}_report_{today}.pdf"
    fpath = os.path.join(REPORTS_DIR, fname)
    s = _get_summary()

    doc = SimpleDocTemplate(fpath, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title2", parent=styles["Title"], textColor=colors.HexColor("#2F5496"))
    elements = []

    elements.append(Paragraph(f"{COMPANY_NAME} — {report_type.title()} Claims Report", title_style))
    elements.append(Paragraph(f"Generated: {today}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    # KPI summary
    kpi_data = [
        ["Total Claims", "Total Claimed", "Avg Claim", "Pass Rate"],
        [f"{s['total_claims']:,}", f"${s['total_claimed']:,.2f}", f"${s['avg_claim']:,.2f}", "—"],
    ]
    kpi_table = Table(kpi_data, colWidths=[180, 180, 180, 180])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 20))

    # Claims by type table
    elements.append(Paragraph("Claims by Type", styles["Heading2"]))
    type_data = [["Type", "Count", "Total Claimed", "Avg", "Approval %"]]
    for t in s["by_type"]:
        type_data.append([
            t["claim_type"], str(t["total_claims"]),
            f"${t['total_claimed']:,.2f}", f"${t['avg_claimed']:,.2f}",
            f"{t['approval_rate_pct'] or 0:.1f}%"
        ])
    type_table = Table(type_data, colWidths=[140, 80, 140, 120, 100])
    type_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    elements.append(type_table)

    doc.build(elements)
    return fpath
